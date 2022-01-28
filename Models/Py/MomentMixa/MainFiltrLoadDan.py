from openpyxl import load_workbook
import numpy as np
import os
import matplotlib.pyplot as plt

import scipy.io as sio

'''
- APP_r – процент нажатия педали
- p_7_INV_1_ActualEngineSpeedFromCAN – обороты двигателя
- AccPed_trqEng_mp – значение момента из карты.

 from  openpyxl import Workbook
'''

wb, ws = dict(), dict()


def find_index(ibegin, iend, names: list):
    cell_range = ws[ibegin:iend]
    count = len(cell_range[0])
    ls = ["0"]
    _ = [ls.append(str(ws.cell(row=1, column=i + 1).value).lower()) for i in range(0, count)]
    names = [str(_z).lower() for _z in names]
    fun1 = lambda x, l: [z for z in l if x in z]
    fun2 = lambda y: fun1(y, ls)[0] if len(fun1(y, ls)) > 0 else '0'
    _ls = [ls.index(fun2(_v)) for _v in names]
    return _ls


def calc(index):
    print(index)
    _lit = ws.cell(row=1, column=index).column_letter
    row_count = ws.max_row
    colC = ws[f'{_lit}2:{_lit}' + str(row_count)]
    cal_app = [z[0].value for z in colC]
    cal_app = [0 if w is None else w for w in cal_app]
    __m = np.array(cal_app, dtype='float32')
    return __m


if __name__ == '__main__':
    wb = load_workbook("E:\Motorcycle\Data\MatLab\Mixail\HVB_pwr2.xlsx", read_only=True)
    _db_name = "HVBMixa"
    __namefile = f"E:\Motorcycle\Data\MatLab\Mixail\{_db_name}.mat"
    print(wb.sheetnames)
    ws = wb.active
#    _ls_name = ["AC_pwrAct"]
#    ls = find_index('A1', 'P1', ['AC_pwrAct' ])
    _ls_name = ["HVB_pwrAct"]
    ls = find_index('A1', 'P1', ['HVB_pwrAct'])

    try:
        _z = ls.index(0)
        print(f" Ошибка в название столбца \n  {ls}")
        os._exit(-1)
    except:
        print(" Столбцы определены !")

    danx = {_ls_name[i]: calc(ls[i]) for i in range(len(_ls_name))}
    # sio.savemat(__namefile, {_db_name:danx})


    
    for key, val in danx.items():
        print(key)
        np.save(f'E:\Motorcycle\Data\Py\Moment/1/{key}', np.array(val))
        np.save(f'E:\Motorcycle\Data\Py\Moment\Filtr/{key}', np.array(val))

    # with open('E:\Motorcycle\Data\Py\Moment\Filtr/HVB.npy', 'rb') as f:
    #     a = np.load(f)

    kkk=1
    # plt.figure()
    # plt.plot(a)
    # plt.show()

    print("  !!!   все  !!")

    k = 1

'''
with open('test.npy', 'wb') as f:
    np.save(f, np.array([1, 2]))
    np.save(f, np.array([1, 3]))
with open('test.npy', 'rb') as f:
    a = np.load(f)
    b = np.load(f)
'''
